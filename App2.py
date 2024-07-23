'''import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import fdb
import openpyxl
from collections import defaultdict
import xlrd

def log_message(message):
    log_text.config(state=tk.NORMAL)
    log_text.insert(tk.END, message + "\n")
    log_text.config(state=tk.DISABLED)
    log_text.yview(tk.END)  # Scroll to the end
    root.update_idletasks()  # Force update of the GUI

def select_database():
    file_path = filedialog.askopenfilename(filetypes=[("Database files", "*.fdb")])
    if file_path:
        db_entry.delete(0, tk.END)
        db_entry.insert(0, file_path)

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        file_entry.delete(0, tk.END)
        file_entry.insert(0, file_path)

def show_confirmation(modifications):
    confirm_window = tk.Toplevel(root)
    confirm_window.title("Confirmar Modificações")

    tk.Label(confirm_window, text="Os seguintes itens serão modificados:").pack(pady=10)

    # Criar o frame para a Treeview e a barra de rolagem
    frame = tk.Frame(confirm_window)
    frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

    # Criar a barra de rolagem vertical
    scrollbar = tk.Scrollbar(frame, orient=tk.VERTICAL)

    # Criar a Treeview
    columns = ('reduzido', 'descricao', 'localizacao_antigo', 'localizacao_novo')
    tree = ttk.Treeview(frame, columns=columns, show='headings', yscrollcommand=scrollbar.set)
    tree.heading('reduzido', text='ITEM_REDUZIDO')
    tree.heading('descricao', text='ITEM_DESCRICAO')
    tree.heading('localizacao_antigo', text='ITEM_LOCALIZACAO ANTIGO')
    tree.heading('localizacao_novo', text='ITEM_LOCALIZACAO NOVO')

    # Configurar a barra de rolagem
    scrollbar.config(command=tree.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    for mod in modifications:
        tree.insert('', tk.END, values=(mod['reduzido'], mod['descricao'], mod['localizacao_antigo'], mod['localizacao_novo']))

    def on_confirm():
        update_database(modifications)
        confirm_window.destroy()

    tk.Button(confirm_window, text="Confirmar", command=on_confirm).pack(side=tk.LEFT, padx=10, pady=10)
    tk.Button(confirm_window, text="Cancelar", command=confirm_window.destroy).pack(side=tk.RIGHT, padx=10, pady=10)

def load_data():
    ipv4_address = ipv4_entry.get()
    port = port_entry.get()
    db_file = db_entry.get()
    file_path = file_entry.get()
    
    # Alterar para a aba de execução
    notebook.select(1)

    if not db_file or not file_path:
        messagebox.showerror("Erro", "Por favor, selecione o arquivo do banco de dados e o arquivo XLSX ou XLS.")
        return

    if ipv4_address and port:
        dsn = f"{ipv4_address}/{port}:{db_file}"
    else:
        dsn = db_file

    try:
        log_message("Conectando ao banco de dados...")
        # Conexão com o banco de dados Firebird
        con = fdb.connect(
            dsn=dsn,
            user='SYSDBA',
            password='masterkey',
            charset='UTF8'
        )
        cursor = con.cursor()
        log_message("Conexão estabelecida com sucesso.")

        # Carregar a planilha
        log_message("Carregando planilha...")
        if file_path.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
        else:  # .xls
            wb = xlrd.open_workbook(file_path)
            sheet = wb.sheet_by_index(0)
        log_message("Planilha carregada com sucesso.")

        # Encontrar índices das colunas relevantes
        if file_path.endswith('.xlsx'):
            col_names = [cell.value for cell in sheet[1]]
        else:
            col_names = [sheet.cell_value(0, col) for col in range(sheet.ncols)]

        rua_idx = col_names.index("Rua")
        posicao_idx = col_names.index("Posição")
        reduzido_idx = col_names.index("Reduzido")

        # Dicionário para acumular localizações por reduzido
        localizacao_dict = defaultdict(list)
        modificacoes = []

        # Iterar pelas linhas da planilha
        log_message("Processando linhas da planilha...")
        if file_path.endswith('.xlsx'):
            for row in sheet.iter_rows(min_row=2):  # Assumindo que a primeira linha é o cabeçalho
                reduzido = row[reduzido_idx].value
                rua = row[rua_idx].value
                posicao = row[posicao_idx].value

                # Garantir que reduzido, rua e posicao não são None
                if reduzido is None or rua is None or posicao is None:
                    continue

                reduzido = str(reduzido).strip()
                localizacao = str(rua).strip() + str(posicao).strip()

                if len(localizacao) > 70:
                    localizacao = localizacao[:70]

                # Acumular localizações no dicionário
                localizacao_dict[reduzido].append(localizacao)
        else:
            for row_idx in range(1, sheet.nrows):  # Assumindo que a primeira linha é o cabeçalho
                reduzido = sheet.cell_value(row_idx, reduzido_idx)
                rua = sheet.cell_value(row_idx, rua_idx)
                posicao = sheet.cell_value(row_idx, posicao_idx)

                # Garantir que reduzido, rua e posicao não são None
                if reduzido is None or rua is None or posicao is None:
                    continue

                reduzido = str(reduzido).strip()
                localizacao = str(rua).strip() + str(posicao).strip()

                if len(localizacao) > 70:
                    localizacao = localizacao[:70]

                # Acumular localizações no dicionário
                localizacao_dict[reduzido].append(localizacao)

        log_message("Linhas processadas. Recuperando descrições dos itens e valores antigos...")
        # Recuperar descrições dos itens e valores antigos
        for reduzido, localizacoes in localizacao_dict.items():
            localizacao_final = ' / '.join(localizacoes)

            cursor.execute("SELECT ITEM_DESCRICAO, ITEM_LOCALIZACAO FROM ITENS WHERE ITEM_REDUZIDO = ? AND EMPRESA_CODIGO = '0001'", (reduzido,))
            result = cursor.fetchone()
            if result:
                descricao = result[0]
                localizacao_antiga = result[1] if result[1] else ""
                modificacoes.append({
                    'reduzido': reduzido,
                    'descricao': descricao,
                    'localizacao_antigo': localizacao_antiga,
                    'localizacao_novo': localizacao_final
                })

        cursor.close()
        con.close()

        log_message("Dados carregados com sucesso. Exibindo confirmação...")
        # Exibir a confirmação
        show_confirmation(modificacoes)

    except Exception as e:
        log_message(f"Erro: {e}")
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

def update_database(modifications):
    ipv4_address = ipv4_entry.get()
    port = port_entry.get()
    db_file = db_entry.get()

    if ipv4_address and port:
        dsn = f"{ipv4_address}/{port}:{db_file}"
    else:
        dsn = db_file

    try:
        log_message("Conectando ao banco de dados para atualização...")
        # Conexão com o banco de dados Firebird2
        con = fdb.connect(
            dsn=dsn,
            user='SYSDBA',
            password='masterkey',
            charset='UTF8'
        )
        cursor = con.cursor()
        log_message("Conexão estabelecida com sucesso.")

        # Atualizar o banco de dados
        for mod in modifications:
            log_message(f"Atualizando ITEM_REDUZIDO: {mod['reduzido']} com ITEM_LOCALIZACAO: {mod['localizacao_novo']}")
            cursor.execute(
                "UPDATE ITENS SET ITEM_LOCALIZACAO = ? WHERE ITEM_REDUZIDO = ? AND EMPRESA_CODIGO = '0001'",
                (mod['localizacao_novo'], mod['reduzido'])
            )

        con.commit()
        cursor.close()
        con.close()
        log_message("Atualização concluída com sucesso.")
        messagebox.showinfo("Sucesso", "Atualização concluída com sucesso.")
    except Exception as e:
        log_message(f"Erro durante a atualização: {e}")
        messagebox.showerror("Erro", f"Ocorreu um erro durante a atualização: {e}")

# Criação da janela principal
root = tk.Tk()
root.title("Atualização de Itens")

# Criação do notebook para as abas
notebook = ttk.Notebook(root)
notebook.pack(pady=10, expand=True)

# Criação da aba de configurações
settings_frame = ttk.Frame(notebook)
notebook.add(settings_frame, text="Configurações")

# Criação do contêiner centralizado
settings_container = ttk.Frame(settings_frame)
settings_container.pack(expand=True, pady=20)

# Configurar o layout do contêiner usando grid com padding
settings_container.grid_columnconfigure(0, weight=1)
settings_container.grid_columnconfigure(1, weight=1)

# Entrada para o host do banco de dados
ipv4_label = ttk.Label(settings_container, text="Host do banco de dados:")
ipv4_label.grid(row=0, column=0, padx=(0, 10), pady=(0, 5), sticky="E")
ipv4_entry = ttk.Entry(settings_container)
ipv4_entry.grid(row=0, column=1, padx=(10, 0), pady=(0, 5), sticky="W")

# Entrada para a porta do Firebird
port_label = ttk.Label(settings_container, text="Porta do Firebird:")
port_label.grid(row=1, column=0, padx=(0, 10), pady=(0, 5), sticky="E")
port_entry = ttk.Entry(settings_container)
port_entry.grid(row=1, column=1, padx=(10, 0), pady=(0, 5), sticky="W")

# Entrada para o arquivo do banco de dados
db_label = ttk.Label(settings_container, text="Arquivo do banco de dados:")
db_label.grid(row=2, column=0, padx=(0, 10), pady=(0, 5), sticky="E")
db_entry = ttk.Entry(settings_container)
db_entry.grid(row=2, column=1, padx=(10, 0), pady=(0, 5), sticky="W")
db_button = ttk.Button(settings_container, text="Selecionar...", command=select_database)
db_button.grid(row=2, column=2, padx=(10, 0), pady=(0, 5))

# Entrada para o arquivo Excel
file_label = ttk.Label(settings_container, text="Arquivo Excel:")
file_label.grid(row=3, column=0, padx=(0, 10), pady=(0, 5), sticky="E")
file_entry = ttk.Entry(settings_container)
file_entry.grid(row=3, column=1, padx=(10, 0), pady=(0, 5), sticky="W")
file_button = ttk.Button(settings_container, text="Selecionar...", command=select_file)
file_button.grid(row=3, column=2, padx=(10, 0), pady=(0, 5))

# Botão para carregar dados
load_button = ttk.Button(settings_container, text="Carregar Dados", command=load_data)
load_button.grid(row=4, column=0, columnspan=3, pady=10)

# Criação da aba de execução
execution_frame = ttk.Frame(notebook)
notebook.add(execution_frame, text="Execução")

log_text = tk.Text(execution_frame, state=tk.DISABLED, wrap=tk.WORD)
log_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

# Iniciar o loop principal do Tkinter
root.mainloop()
'''

import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import fdb
import openpyxl
from collections import defaultdict
import xlrd
import threading

def log_message(message):
    log_text.config(state=tk.NORMAL)
    log_text.insert(tk.END, message + "\n")
    log_text.config(state=tk.DISABLED)
    log_text.yview(tk.END)  # Scroll to the end
    root.update_idletasks()  # Force update of the GUI

def select_database():
    file_path = filedialog.askopenfilename(filetypes=[("Database files", "*.fdb")])
    if file_path:
        db_entry.delete(0, tk.END)
        db_entry.insert(0, file_path)

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        file_entry.delete(0, tk.END)
        file_entry.insert(0, file_path)

def show_confirmation(modifications):
    confirm_window = tk.Toplevel(root)
    confirm_window.title("Confirmar Modificações")

    tk.Label(confirm_window, text="Os seguintes itens serão modificados:").pack(pady=10)

    frame = tk.Frame(confirm_window)
    frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

    scrollbar = tk.Scrollbar(frame, orient=tk.VERTICAL)
    columns = ('reduzido', 'descricao', 'localizacao_antigo', 'localizacao_novo')
    tree = ttk.Treeview(frame, columns=columns, show='headings', yscrollcommand=scrollbar.set)
    tree.heading('reduzido', text='ITEM_REDUZIDO')
    tree.heading('descricao', text='ITEM_DESCRICAO')
    tree.heading('localizacao_antigo', text='ITEM_LOCALIZACAO ANTIGO')
    tree.heading('localizacao_novo', text='ITEM_LOCALIZACAO NOVO')

    scrollbar.config(command=tree.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    for mod in modifications:
        tree.insert('', tk.END, values=(mod['reduzido'], mod['descricao'], mod['localizacao_antigo'], mod['localizacao_novo']))

    def on_confirm():
        threading.Thread(target=update_database, args=(modifications,)).start()
        confirm_window.destroy()

    tk.Button(confirm_window, text="Confirmar", command=on_confirm).pack(side=tk.LEFT, padx=10, pady=10)
    tk.Button(confirm_window, text="Cancelar", command=confirm_window.destroy).pack(side=tk.RIGHT, padx=10, pady=10)

def load_data():
    threading.Thread(target=load_data_thread).start()

def load_data_thread():
    ipv4_address = ipv4_entry.get()
    port = port_entry.get()
    db_file = db_entry.get()
    file_path = file_entry.get()

    notebook.select(1)

    if not db_file or not file_path:
        messagebox.showerror("Erro", "Por favor, selecione o arquivo do banco de dados e o arquivo XLSX ou XLS.")
        return

    if ipv4_address and port:
        dsn = f"{ipv4_address}/{port}:{db_file}"
    else:
        dsn = db_file

    try:
        log_message("Conectando ao banco de dados...")
        con = fdb.connect(dsn=dsn, user='SYSDBA', password='masterkey', charset='UTF8')
        cursor = con.cursor()
        log_message("Conexão estabelecida com sucesso.")

        log_message("Carregando planilha...")
        if file_path.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
        else:
            wb = xlrd.open_workbook(file_path)
            sheet = wb.sheet_by_index(0)
        log_message("Planilha carregada com sucesso.")

        if file_path.endswith('.xlsx'):
            col_names = [cell.value for cell in sheet[1]]
        else:
            col_names = [sheet.cell_value(0, col) for col in range(sheet.ncols)]

        rua_idx = col_names.index("Rua")
        posicao_idx = col_names.index("Posição")
        reduzido_idx = col_names.index("Reduzido")

        localizacao_dict = defaultdict(list)
        modificacoes = []

        log_message("Processando linhas da planilha...")
        if file_path.endswith('.xlsx'):
            for row in sheet.iter_rows(min_row=2):
                reduzido = row[reduzido_idx].value
                rua = row[rua_idx].value
                posicao = row[posicao_idx].value

                if reduzido is None or rua is None or posicao is None:
                    continue

                reduzido = str(reduzido).strip()
                localizacao = str(rua).strip() + str(posicao).strip()

                if len(localizacao) > 70:
                    localizacao = localizacao[:70]

                localizacao_dict[reduzido].append(localizacao)
        else:
            for row_idx in range(1, sheet.nrows):
                reduzido = sheet.cell_value(row_idx, reduzido_idx)
                rua = sheet.cell_value(row_idx, rua_idx)
                posicao = sheet.cell_value(row_idx, posicao_idx)

                if reduzido is None or rua is None or posicao is None:
                    continue

                reduzido = str(reduzido).strip()
                localizacao = str(rua).strip() + str(posicao).strip()

                if len(localizacao) > 70:
                    localizacao = localizacao[:70]

                localizacao_dict[reduzido].append(localizacao)

        log_message("Linhas processadas. Recuperando descrições dos itens e valores antigos...")
        for reduzido, localizacoes in localizacao_dict.items():
            localizacao_final = ' / '.join(localizacoes)
            cursor.execute("SELECT ITEM_DESCRICAO, ITEM_LOCALIZACAO FROM ITENS WHERE ITEM_REDUZIDO = ? AND EMPRESA_CODIGO = '0001'", (reduzido,))
            result = cursor.fetchone()
            if result:
                descricao = result[0]
                localizacao_antiga = result[1] if result[1] else ""
                modificacoes.append({
                    'reduzido': reduzido,
                    'descricao': descricao,
                    'localizacao_antigo': localizacao_antiga,
                    'localizacao_novo': localizacao_final
                })

        cursor.close()
        con.close()

        log_message("Dados carregados com sucesso. Exibindo confirmação...")
        show_confirmation(modificacoes)

    except Exception as e:
        log_message(f"Erro: {e}")
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

def update_database(modifications):
    ipv4_address = ipv4_entry.get()
    port = port_entry.get()
    db_file = db_entry.get()

    if ipv4_address and port:
        dsn = f"{ipv4_address}/{port}:{db_file}"
    else:
        dsn = db_file

    def update_chunk(chunk):
        try:
            log_message("Conectando ao banco de dados para atualização...")
            con = fdb.connect(dsn=dsn, user='SYSDBA', password='masterkey', charset='UTF8')
            cursor = con.cursor()
            log_message("Conexão estabelecida com sucesso.")

            for mod in chunk:
                log_message(f"Atualizando ITEM_REDUZIDO: {mod['reduzido']} com ITEM_LOCALIZACAO: {mod['localizacao_novo']}")
                cursor.execute(
                    "UPDATE ITENS SET ITEM_LOCALIZACAO = ? WHERE ITEM_REDUZIDO = ? AND EMPRESA_CODIGO = '0001'",
                    (mod['localizacao_novo'], mod['reduzido'])
                )
            con.commit()
            cursor.close()
            con.close()
            log_message("Chunk atualizado com sucesso.")

        except Exception as e:
            log_message(f"Erro ao atualizar chunk: {e}")

    try:
        chunk_size = 100
        for i in range(0, len(modifications), chunk_size):
            chunk = modifications[i:i+chunk_size]
            threading.Thread(target=update_chunk, args=(chunk,)).start()

        log_message("Atualizações iniciadas.")
    except Exception as e:
        log_message(f"Erro: {e}")
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

root = tk.Tk()
root.title("Atualização de Banco de Dados Firebird")
root.geometry("650x300")

notebook = ttk.Notebook(root)
notebook.pack(pady=10, expand=True, fill=tk.BOTH)

aba_config = ttk.Frame(notebook)
aba_execucao = ttk.Frame(notebook)
notebook.add(aba_config, text='Configurações')
notebook.add(aba_execucao, text='Execução')

frame1 = ttk.Frame(aba_config)
frame1.pack(pady=20, padx=20, fill='x')

ttk.Label(frame1, text="Endereço do banco de dados (IPv4):").grid(row=0, column=0, sticky='w')
ipv4_entry = ttk.Entry(frame1)
ipv4_entry.grid(row=0, column=1)

ttk.Label(frame1, text="Porta do banco de dados:").grid(row=1, column=0, sticky='w')
port_entry = ttk.Entry(frame1)
port_entry.grid(row=1, column=1)

ttk.Label(frame1, text="Arquivo do banco de dados:").grid(row=2, column=0, sticky='w')
db_entry = ttk.Entry(frame1, width=50)
db_entry.grid(row=2, column=1)
ttk.Button(frame1, text="Selecionar", command=select_database).grid(row=2, column=2, padx=10)

ttk.Label(frame1, text="Arquivo XLSX ou XLS:").grid(row=3, column=0, sticky='w')
file_entry = ttk.Entry(frame1, width=50)
file_entry.grid(row=3, column=1)
ttk.Button(frame1, text="Selecionar", command=select_file).grid(row=3, column=2, padx=10)

ttk.Button(frame1, text="Carregar Dados", command=load_data).grid(row=4, columnspan=3, pady=10)

frame2 = ttk.Frame(aba_execucao)
frame2.pack(pady=20, padx=20, fill='x')

ttk.Label(frame2, text="Log de Execução:").pack(anchor='w')
log_text = tk.Text(frame2, wrap=tk.WORD, height=10, state=tk.DISABLED)
log_text.pack(fill=tk.BOTH, padx=10, pady=10, expand=True)

root.mainloop()
