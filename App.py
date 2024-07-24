import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import fdb
import openpyxl
from collections import defaultdict
import xlrd

# Função para exibir mensagens no log
def log_message(message):
    log_text.config(state=tk.NORMAL) # Habilitar a edição do texto
    log_text.insert(tk.END, message + "\n") # Adicionar a mensagem ao final do texto
    log_text.config(state=tk.DISABLED) # Desabilitar a edição do texto
    log_text.yview(tk.END)  # Deslocar a barra de rolagem para o final
    root.update_idletasks()  # Forçar a atualização da interface

# Funções para selecionar arquivos
def select_database():
    file_path = filedialog.askopenfilename(filetypes=[("Database files", "*.fdb")])
    if file_path:
        db_entry.delete(0, tk.END)
        db_entry.insert(0, file_path)

# Função para selecionar arquivos
def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        file_entry.delete(0, tk.END)
        file_entry.insert(0, file_path)

# Função para exibir a confirmação
def show_confirmation(modifications):
    confirm_window = tk.Toplevel(root)
    confirm_window.title("Confirmar Modificações")

    tk.Label(confirm_window, text="Os seguintes itens serão modificados:").pack(pady=10)

    # Criar o frame para a Treeview e a barra de rolagem
    frame = tk.Frame(confirm_window)
    frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

    # Criar a barra de rolagem vertical
    scrollbar = tk.Scrollbar(frame, orient=tk.VERTICAL)

    # Criar a Treeview
    columns = ('reduzido', 'descricao', 'localizacao_antigo', 'localizacao_novo')
    tree = ttk.Treeview(frame, columns=columns, show='headings', yscrollcommand=scrollbar.set)
    tree.heading('reduzido', text='ITEM_REDUZIDO')
    tree.heading('descricao', text='ITEM_DESCRICAO')
    tree.heading('localizacao_antigo', text='ITEM_LOCALIZACAO ANTIGO')
    tree.heading('localizacao_novo', text='ITEM_LOCALIZACAO NOVO')

    # Configurar a barra de rolagem
    scrollbar.config(command=tree.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    for mod in modifications:
        tree.insert('', tk.END, values=(mod['reduzido'], mod['descricao'], mod['localizacao_antigo'], mod['localizacao_novo']))

    def on_confirm():
        update_database(modifications)
        confirm_window.destroy()

    tk.Button(confirm_window, text="Confirmar", command=on_confirm).pack(side=tk.LEFT, padx=10, pady=10)
    tk.Button(confirm_window, text="Cancelar", command=confirm_window.destroy).pack(side=tk.RIGHT, padx=10, pady=10)

# Função para carregar os dados
def load_data():
    ipv4_address = ipv4_entry.get()
    port = port_entry.get()
    db_file = db_entry.get()
    file_path = file_entry.get()
    
    #Alterar para a aba de execução
    notebook.select(1)

    if not db_file or not file_path:
        messagebox.showerror("Erro", "Por favor, selecione o arquivo do banco de dados e o arquivo XLSX ou XLS.")
        return

    if ipv4_address and port:
        dsn = f"{ipv4_address}/{port}:{db_file}"
    else:
        dsn = db_file

    try:
        log_message("Conectando ao banco de dados...")
        # Conexão com o banco de dados Firebird
        con = fdb.connect(
            dsn=dsn,
            user='SYSDBA',
            password='masterkey',
            charset='UTF8'
        )
        cursor = con.cursor()
        log_message("Conexão estabelecida com sucesso.")

        # Carregar a planilha
        log_message("Carregando planilha...")
        if file_path.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
        else:  # .xls
            wb = xlrd.open_workbook(file_path)
            sheet = wb.sheet_by_index(0)
        log_message("Planilha carregada com sucesso.")

        # Encontrar índices das colunas relevantes
        if file_path.endswith('.xlsx'):
            col_names = [cell.value for cell in sheet[1]]
        else:
            col_names = [sheet.cell_value(0, col) for col in range(sheet.ncols)]

        rua_idx = col_names.index("Rua")
        posicao_idx = col_names.index("Posição")
        reduzido_idx = col_names.index("Reduzido")

        # Dicionário para acumular localizações por reduzido
        localizacao_dict = defaultdict(list)
        modificacoes = []

        # Iterar pelas linhas da planilha
        log_message("Processando linhas da planilha...")
        if file_path.endswith('.xlsx'):
            for row in sheet.iter_rows(min_row=2):  # Assumindo que a primeira linha é o cabeçalho
                reduzido = row[reduzido_idx].value
                rua = row[rua_idx].value
                posicao = row[posicao_idx].value

                # Garantir que reduzido, rua e posicao não são None
                if reduzido is None or rua is None or posicao is None:
                    continue

                reduzido = str(reduzido).strip()
                localizacao = str(rua).strip() + str(posicao).strip()

                if len(localizacao) > 70:
                    localizacao = localizacao[:70]

                # Acumular localizações no dicionário
                localizacao_dict[reduzido].append(localizacao)
        else:
            for row_idx in range(1, sheet.nrows):  # Assumindo que a primeira linha é o cabeçalho
                reduzido = sheet.cell_value(row_idx, reduzido_idx)
                rua = sheet.cell_value(row_idx, rua_idx)
                posicao = sheet.cell_value(row_idx, posicao_idx)

                # Garantir que reduzido, rua e posicao não são None
                if reduzido is None or rua is None or posicao is None:
                    continue
                
                # Convertendo para string e removendo espaços em branco
                reduzido = str(reduzido).strip()
                localizacao = str(rua).strip() + str(posicao).strip()

                if len(localizacao) > 70:
                    localizacao = localizacao[:70]

                # Acumular localizações no dicionário
                localizacao_dict[reduzido].append(localizacao)

        log_message("Linhas processadas. Recuperando descrições dos itens e valores antigos...")
        # Recuperar descrições dos itens e valores antigos
        for reduzido, localizacoes in localizacao_dict.items():
            localizacao_final = ' / '.join(localizacoes)

            cursor.execute("SELECT ITEM_DESCRICAO, ITEM_LOCALIZACAO FROM ITENS WHERE ITEM_REDUZIDO = ? AND EMPRESA_CODIGO = '0001'", (reduzido,))
            result = cursor.fetchone()
            if result:
                descricao = result[0]
                localizacao_antiga = result[1] if result[1] else ""
                modificacoes.append({
                    'reduzido': reduzido,
                    'descricao': descricao,
                    'localizacao_antigo': localizacao_antiga,
                    'localizacao_novo': localizacao_final
                })

        cursor.close()
        con.close()

        log_message("Dados carregados com sucesso. Exibindo confirmação...")
        # Exibir a confirmação
        show_confirmation(modificacoes)

    except Exception as e:
        log_message(f"Erro: {e}")
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

# Função para atualizar o banco de dados
def update_database(modifications):
    ipv4_address = ipv4_entry.get()
    port = port_entry.get()
    db_file = db_entry.get()

    if ipv4_address and port:
        dsn = f"{ipv4_address}/{port}:{db_file}"
    else:
        dsn = db_file

    try:
        log_message("Conectando ao banco de dados para atualização...")
        # Conexão com o banco de dados Firebird2
        con = fdb.connect(
            dsn=dsn,
            user='SYSDBA',
            password='masterkey',
            charset='UTF8'
        )
        cursor = con.cursor()
        log_message("Conexão estabelecida com sucesso.")

        # Atualizar o banco de dados
        for mod in modifications:
            log_message(f"Atualizando ITEM_REDUZIDO: {mod['reduzido']} com ITEM_LOCALIZACAO: {mod['localizacao_novo']}")
            cursor.execute(
                "UPDATE ITENS SET ITEM_LOCALIZACAO = ? WHERE ITEM_REDUZIDO = ? AND EMPRESA_CODIGO = '0001'",
                (mod['localizacao_novo'], mod['reduzido'])
            )

        con.commit()
        cursor.close()
        con.close()
        log_message("Atualização concluída com sucesso.")
        messagebox.showinfo("Sucesso", "Atualização concluída com sucesso.")

    except Exception as e:
        log_message(f"Erro: {e}")
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

# Configuração da interface Tkinter
root = tk.Tk()
root.title("Atualização de Banco de Dados Firebird")
root.geometry("650x300")

# Configuração do notebook
notebook = ttk.Notebook(root)
notebook.pack(pady=10, expand=True, fill=tk.BOTH) #, fill=tk.BOTH

# Frames para cada aba
aba_config = ttk.Frame(notebook)
aba_execucao = ttk.Frame(notebook)
notebook.add(aba_config, text='Configurações')
notebook.add(aba_execucao, text='Execução')

# Usando o pack_configure para preencher totalemente a janela com o notebook da seguinte forma:
notebook.pack_configure(fill=tk.BOTH, expand=True)

# Widgets da aba de Configurações
frame1 = ttk.Frame(aba_config)
frame1.pack(pady=20, padx=20) #, fill='x'

# Configuração das labels e entries usando grid
ttk.Label(frame1, text="Host do banco de dados:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
ipv4_entry = ttk.Entry(frame1, width=25)
ipv4_entry.grid(row=0, column=1, padx=(0, 60), pady=5)

ttk.Label(frame1, text="Porta do Firebird:").grid(row=0, column=2, sticky=tk.W, padx=(0, 10))
port_entry = ttk.Entry(frame1, width=12)
port_entry.grid(row=0, column=3, padx=(0, 20), pady=5)

# Configuração do 2º frame
frame2 = ttk.Frame(aba_config)
frame2.pack(pady=20, padx=20) #, fill='x'

ttk.Label(frame2, text="Arquivo do Banco de Dados:").grid(row=0, column=0, sticky=tk.W)
db_entry = ttk.Entry(frame2, width=50)
db_entry.grid(row=0, column=1, padx=0, pady=5)
db_button = ttk.Button(frame2, text="Selecionar", command=select_database)
db_button.grid(row=0, column=2, padx=10)

# Adicionar um Label com o asterisco vermelho
ttk.Label(frame2, text="*", foreground="red").grid(row=0, column=3, sticky=tk.W, padx=(0, 5))

ttk.Label(frame2, text="Arquivo XLSX ou XLS:").grid(row=1, column=0, sticky=tk.W)
file_entry = ttk.Entry(frame2, width=50)
file_entry.grid(row=1, column=1, padx=0, pady=5)
file_button = ttk.Button(frame2, text="Selecionar", command=select_file)
file_button.grid(row=1, column=2, padx=10)

# Adicionar um Label com o asterisco vermelho
ttk.Label(frame2, text="*", foreground="red").grid(row=1, column=3, sticky=tk.W, padx=(0, 5))

# Botão para carregar os dados
load_button = ttk.Button(aba_config, text="Carregar Dados", command=load_data)
load_button.pack(pady=20)

# Widgets da aba de Execução
log_text = tk.Text(aba_execucao, wrap=tk.WORD, state=tk.DISABLED)
log_text.pack(pady=20, padx=20, fill=tk.BOTH, expand=True)

# Barra de rolagem para o log
scrollbar = ttk.Scrollbar(log_text, command=log_text.yview)
log_text['yscrollcommand'] = scrollbar.set
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

root.mainloop()

'''
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import fdb
import openpyxl
from collections import defaultdict
import xlrd
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.mime.text import MIMEText

# Função para exibir mensagens no log
def log_message(message):
    log_text.config(state=tk.NORMAL) # Habilitar a edição do texto
    log_text.insert(tk.END, message + "\n") # Adicionar a mensagem ao final do texto
    log_text.config(state=tk.DISABLED) # Desabilitar a edição do texto
    log_text.yview(tk.END)  # Deslocar a barra de rolagem para o final
    root.update_idletasks()  # Forçar a atualização da interface

# Funções para selecionar arquivos
def select_database():
    file_path = filedialog.askopenfilename(filetypes=[("Database files", "*.fdb")])
    if file_path:
        db_entry.delete(0, tk.END)
        db_entry.insert(0, file_path)

# Função para selecionar arquivos
def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        file_entry.delete(0, tk.END)
        file_entry.insert(0, file_path)

# Função para exibir a confirmação
def show_confirmation(modifications):
    confirm_window = tk.Toplevel(root)
    confirm_window.title("Confirmar Modificações")

    tk.Label(confirm_window, text="Os seguintes itens serão modificados:").pack(pady=10)

    # Criar o frame para a Treeview e a barra de rolagem
    frame = tk.Frame(confirm_window)
    frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

    # Criar a barra de rolagem vertical
    scrollbar = tk.Scrollbar(frame, orient=tk.VERTICAL)

    # Criar a Treeview
    columns = ('reduzido', 'descricao', 'localizacao_antigo', 'localizacao_novo')
    tree = ttk.Treeview(frame, columns=columns, show='headings', yscrollcommand=scrollbar.set)
    tree.heading('reduzido', text='ITEM_REDUZIDO')
    tree.heading('descricao', text='ITEM_DESCRICAO')
    tree.heading('localizacao_antigo', text='ITEM_LOCALIZACAO ANTIGO')
    tree.heading('localizacao_novo', text='ITEM_LOCALIZACAO NOVO')

    # Configurar a barra de rolagem
    scrollbar.config(command=tree.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    for mod in modifications:
        tree.insert('', tk.END, values=(mod['reduzido'], mod['descricao'], mod['localizacao_antigo'], mod['localizacao_novo']))

    def on_confirm():
        update_database(modifications)
        confirm_window.destroy()

    tk.Button(confirm_window, text="Confirmar", command=on_confirm).pack(side=tk.LEFT, padx=10, pady=10)
    tk.Button(confirm_window, text="Cancelar", command=confirm_window.destroy).pack(side=tk.RIGHT, padx=10, pady=10)

    # Adicionar botão para enviar por email
    tk.Button(confirm_window, text="Enviar por email", command=lambda: send_email_window(modifications)).pack(pady=10)

# Função para carregar os dados
def load_data():
    ipv4_address = ipv4_entry.get()
    port = port_entry.get()
    db_file = db_entry.get()
    file_path = file_entry.get()
    
    #Alterar para a aba de execução
    notebook.select(1)

    if not db_file or not file_path:
        messagebox.showerror("Erro", "Por favor, selecione o arquivo do banco de dados e o arquivo XLSX ou XLS.")
        return

    if ipv4_address and port:
        dsn = f"{ipv4_address}/{port}:{db_file}"
    else:
        dsn = db_file

    try:
        log_message("Conectando ao banco de dados...")
        # Conexão com o banco de dados Firebird
        con = fdb.connect(
            dsn=dsn,
            user='SYSDBA',
            password='masterkey',
            charset='UTF8'
        )
        cursor = con.cursor()
        log_message("Conexão estabelecida com sucesso.")

        # Carregar a planilha
        log_message("Carregando planilha...")
        if file_path.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
        else:  # .xls
            wb = xlrd.open_workbook(file_path)
            sheet = wb.sheet_by_index(0)
        log_message("Planilha carregada com sucesso.")

        # Encontrar índices das colunas relevantes
        if file_path.endswith('.xlsx'):
            col_names = [cell.value for cell in sheet[1]]
        else:
            col_names = [sheet.cell_value(0, col) for col in range(sheet.ncols)]

        rua_idx = col_names.index("Rua")
        posicao_idx = col_names.index("Posição")
        reduzido_idx = col_names.index("Reduzido")

        # Dicionário para acumular localizações por reduzido
        localizacao_dict = defaultdict(list)
        modificacoes = []

        # Iterar pelas linhas da planilha
        log_message("Processando linhas da planilha...")
        if file_path.endswith('.xlsx'):
            for row in sheet.iter_rows(min_row=2):  # Assumindo que a primeira linha é o cabeçalho
                reduzido = row[reduzido_idx].value
                rua = row[rua_idx].value
                posicao = row[posicao_idx].value

                # Garantir que reduzido, rua e posicao não são None
                if reduzido is None or rua is None or posicao is None:
                    continue

                reduzido = str(reduzido).strip()
                localizacao = str(rua).strip() + str(posicao).strip()

                if len(localizacao) > 70:
                    localizacao = localizacao[:70]

                # Acumular localizações no dicionário
                localizacao_dict[reduzido].append(localizacao)
        else:
            for row_idx in range(1, sheet.nrows):  # Assumindo que a primeira linha é o cabeçalho
                reduzido = sheet.cell_value(row_idx, reduzido_idx)
                rua = sheet.cell_value(row_idx, rua_idx)
                posicao = sheet.cell_value(row_idx, posicao_idx)

                # Garantir que reduzido, rua e posicao não são None
                if reduzido is None or rua is None or posicao is None:
                    continue
                
                # Convertendo para string e removendo espaços em branco
                reduzido = str(reduzido).strip()
                localizacao = str(rua).strip() + str(posicao).strip()

                if len(localizacao) > 70:
                    localizacao = localizacao[:70]

                # Acumular localizações no dicionário
                localizacao_dict[reduzido].append(localizacao)

        log_message("Linhas processadas. Recuperando descrições dos itens e valores antigos...")
        # Recuperar descrições dos itens e valores antigos
        for reduzido, localizacoes in localizacao_dict.items():
            localizacao_final = ' / '.join(localizacoes)

            cursor.execute("SELECT ITEM_DESCRICAO, ITEM_LOCALIZACAO FROM ITENS WHERE ITEM_REDUZIDO = ? AND EMPRESA_CODIGO = '0001'", (reduzido,))
            result = cursor.fetchone()
            if result:
                descricao = result[0]
                localizacao_antiga = result[1] if result[1] else ""
                modificacoes.append({
                    'reduzido': reduzido,
                    'descricao': descricao,
                    'localizacao_antigo': localizacao_antiga,
                    'localizacao_novo': localizacao_final
                })

        cursor.close()
        con.close()

        log_message("Dados carregados com sucesso. Exibindo confirmação...")
        # Exibir a confirmação
        show_confirmation(modificacoes)

    except Exception as e:
        log_message(f"Erro: {e}")
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

# Função para atualizar o banco de dados
def update_database(modifications):
    ipv4_address = ipv4_entry.get()
    port = port_entry.get()
    db_file = db_entry.get()

    if ipv4_address and port:
        dsn = f"{ipv4_address}/{port}:{db_file}"
    else:
        dsn = db_file

    try:
        log_message("Conectando ao banco de dados para atualização...")
        # Conexão com o banco de dados Firebird2
        con = fdb.connect(
            dsn=dsn,
            user='SYSDBA',
            password='masterkey',
            charset='UTF8'
        )
        cursor = con.cursor()
        log_message("Conexão estabelecida com sucesso.")

        # Atualizar o banco de dados
        for mod in modifications:
            log_message(f"Atualizando ITEM_REDUZIDO: {mod['reduzido']} com ITEM_LOCALIZACAO: {mod['localizacao_novo']}")
            cursor.execute(
                "UPDATE ITENS SET ITEM_LOCALIZACAO = ? WHERE ITEM_REDUZIDO = ? AND EMPRESA_CODIGO = '0001'",
                (mod['localizacao_novo'], mod['reduzido'])
            )

        con.commit()
        cursor.close()
        con.close()
        log_message("Atualização concluída com sucesso.")
        messagebox.showinfo("Sucesso", "Atualização concluída com sucesso.")

    except Exception as e:
        log_message(f"Erro: {e}")
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

# Função para gerar o arquivo XLSX
def generate_xlsx(modifications, filepath="modifications.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['ITEM_REDUZIDO', 'ITEM_DESCRICAO', 'ITEM_LOCALIZACAO ANTIGO', 'ITEM_LOCALIZACAO NOVO'])
    for mod in modifications:
        ws.append([mod['reduzido'], mod['descricao'], mod['localizacao_antigo'], mod['localizacao_novo']])
    wb.save(filepath)
    return filepath

# Função para enviar email
def send_email(recipient, filepath):
    try:
        sender = "SEU_EMAIL"
        password = "SUA_SENHA_DO_GMAIL"  # Substitua pela sua senha

        msg = MIMEMultipart()
        msg['From'] = sender
        msg['To'] = recipient
        msg['Subject'] = "Modificações no Banco de Dados"

        body = "Segue em anexo o arquivo com as modificações realizadas."
        msg.attach(MIMEText(body, 'plain'))

        attachment = open(filepath, "rb")
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f"attachment; filename= {filepath}")

        msg.attach(part)

        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(sender, password)
        text = msg.as_string()
        server.sendmail(sender, recipient, text)
        server.quit()
        log_message("Email enviado com sucesso.")
    except Exception as e:
        log_message(f"Erro ao enviar email: {e}")
        messagebox.showerror("Erro", f"Erro ao enviar email: {e}")

# Função para abrir a janela de envio de email
def send_email_window(modifications):
    email_window = tk.Toplevel(root)
    email_window.title("Enviar por Email")
    tk.Label(email_window, text="Digite o email do destinatário:").pack(pady=10)
    email_entry = tk.Entry(email_window, width=50)
    email_entry.pack(pady=10)
    def on_send():
        recipient = email_entry.get()
        if recipient:
            filepath = generate_xlsx(modifications)
            send_email(recipient, filepath)
            email_window.destroy()
        else:
            messagebox.showerror("Erro", "Por favor, digite um email válido.")
    tk.Button(email_window, text="Enviar", command=on_send).pack(pady=10)

# Configuração da interface Tkinter
root = tk.Tk()
root.title("Atualização de Banco de Dados Firebird")
root.geometry("650x300")

# Configuração do notebook
notebook = ttk.Notebook(root)
notebook.pack(pady=10, expand=True, fill=tk.BOTH) #, fill=tk.BOTH

# Frames para cada aba
aba_config = ttk.Frame(notebook)
aba_execucao = ttk.Frame(notebook)
notebook.add(aba_config, text='Configurações')
notebook.add(aba_execucao, text='Execução')

# Usando o pack_configure para preencher totalemente a janela com o notebook da seguinte forma:
notebook.pack_configure(fill=tk.BOTH, expand=True)

# Widgets da aba de Configurações
frame1 = ttk.Frame(aba_config)
frame1.pack(pady=20, padx=20) #, fill='x'

# Configuração das labels e entries usando grid
ttk.Label(frame1, text="Host do banco de dados:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
ipv4_entry = ttk.Entry(frame1, width=25)
ipv4_entry.grid(row=0, column=1, padx=(0, 60), pady=5)

ttk.Label(frame1, text="Porta do Firebird:").grid(row=0, column=2, sticky=tk.W, padx=(0, 10))
port_entry = ttk.Entry(frame1, width=12)
port_entry.grid(row=0, column=3, padx=(0, 20), pady=5)

# Configuração do 2º frame
frame2 = ttk.Frame(aba_config)
frame2.pack(pady=20, padx=20) #, fill='x'

ttk.Label(frame2, text="Arquivo do Banco de Dados:").grid(row=0, column=0, sticky=tk.W)
db_entry = ttk.Entry(frame2, width=50)
db_entry.grid(row=0, column=1, padx=0, pady=5)
db_button = ttk.Button(frame2, text="Selecionar", command=select_database)
db_button.grid(row=0, column=2, padx=10)

# Adicionar um Label com o asterisco vermelho
ttk.Label(frame2, text="*", foreground="red").grid(row=0, column=3, sticky=tk.W, padx=(0, 5))

ttk.Label(frame2, text="Arquivo XLSX ou XLS:").grid(row=1, column=0, sticky=tk.W)
file_entry = ttk.Entry(frame2, width=50)
file_entry.grid(row=1, column=1, padx=0, pady=5)
file_button = ttk.Button(frame2, text="Selecionar", command=select_file)
file_button.grid(row=1, column=2, padx=10)

# Adicionar um Label com o asterisco vermelho
ttk.Label(frame2, text="*", foreground="red").grid(row=1, column=3, sticky=tk.W, padx=(0, 5))

# Botão para carregar os dados
load_button = ttk.Button(aba_config, text="Carregar Dados", command=load_data)
load_button.pack(pady=20)

# Widgets da aba de Execução
log_text = tk.Text(aba_execucao, wrap=tk.WORD, state=tk.DISABLED)
log_text.pack(pady=20, padx=20, fill=tk.BOTH, expand=True)

# Barra de rolagem para o log
scrollbar = ttk.Scrollbar(log_text, command=log_text.yview)
log_text['yscrollcommand'] = scrollbar.set
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

root.mainloop()
'''